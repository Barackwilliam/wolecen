from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.core.paginator import Paginator
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST

from .models import PaymentRequest, RequestDocument, Comment, PaymentCategory
from .forms import PaymentRequestForm, CommentForm, RetirementForm
from apps.accounts.models import Department
from apps.approvals.services import ApprovalService
from apps.audit.models import AuditLog

import io


def get_client_ip(request):
    x_fwd = request.META.get('HTTP_X_FORWARDED_FOR')
    return x_fwd.split(',')[0].strip() if x_fwd else request.META.get('REMOTE_ADDR')


# ─── DASHBOARD ────────────────────────────────────────────────
@login_required
def dashboard_view(request):
    user = request.user
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Base queryset based on role
    if user.is_system_admin or user.is_auditor:
        base_qs = PaymentRequest.objects.all()
    elif user.is_finance:
        base_qs = PaymentRequest.objects.all()
    elif user.is_reviewer1 or user.is_reviewer2:
        base_qs = PaymentRequest.objects.all()
    else:
        base_qs = PaymentRequest.objects.filter(requester=user)

    # Stats
    stats = {
        'total': base_qs.count(),
        'pending': base_qs.filter(status__in=[
            PaymentRequest.Status.SUBMITTED,
            PaymentRequest.Status.APPROVED_L1,
            PaymentRequest.Status.PROCESSING,
        ]).count(),
        'paid': base_qs.filter(status__in=[PaymentRequest.Status.PAID, PaymentRequest.Status.CLOSED]).count(),
        'retirement_pending': base_qs.filter(status=PaymentRequest.Status.PAID, has_retirement=True).count(),
        'total_amount_month': base_qs.filter(created_at__gte=month_start).aggregate(
            t=Sum('amount_requested'))['t'] or 0,
        'total_approved_month': base_qs.filter(
            created_at__gte=month_start, amount_approved__isnull=False
        ).aggregate(t=Sum('amount_approved'))['t'] or 0,
        'total_paid_month': base_qs.filter(
            paid_at__gte=month_start
        ).aggregate(t=Sum('amount_approved'))['t'] or 0,
    }

    # Recent requests
    recent_requests = base_qs.select_related('requester', 'department').order_by('-created_at')[:8]

    # Action queue for this user
    action_queue = []
    if user.can_approve_level1:
        action_queue += list(PaymentRequest.objects.filter(
            status=PaymentRequest.Status.SUBMITTED
        ).select_related('requester', 'department').order_by('-priority', '-created_at')[:5])
    if user.can_approve_level2:
        action_queue += list(PaymentRequest.objects.filter(
            status=PaymentRequest.Status.APPROVED_L1
        ).select_related('requester', 'department').order_by('-priority', '-created_at')[:5])
    if user.can_process_payment:
        action_queue += list(PaymentRequest.objects.filter(
            status__in=[PaymentRequest.Status.APPROVED_L2, PaymentRequest.Status.PROCESSING]
        ).select_related('requester', 'department').order_by('-created_at')[:5])
    if user.can_verify_retirement:
        action_queue += list(PaymentRequest.objects.filter(
            status=PaymentRequest.Status.RETIREMENT_SUBMITTED
        ).select_related('requester').order_by('-retired_at')[:5])

    # Recent activity
    from apps.approvals.models import ApprovalAction
    recent_actions = ApprovalAction.objects.select_related('actor', 'request').order_by('-timestamp')[:10]

    return render(request, 'dashboard/home.html', {
        'stats': stats,
        'recent_requests': recent_requests,
        'action_queue': action_queue[:8],
        'recent_actions': recent_actions,
    })


# ─── PAYMENT REQUESTS ─────────────────────────────────────────
@login_required
def payment_list_view(request):
    user = request.user

    if user.is_system_admin or user.is_auditor or user.is_finance or user.is_reviewer1 or user.is_reviewer2:
        qs = PaymentRequest.objects.all()
    else:
        qs = PaymentRequest.objects.filter(requester=user)

    qs = qs.select_related('requester', 'department', 'category')

    # Filters
    q = request.GET.get('q', '').strip()
    status = request.GET.get('status', '')
    dept_id = request.GET.get('department', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    if q:
        qs = qs.filter(
            Q(request_number__icontains=q) | Q(title__icontains=q) |
            Q(requester__first_name__icontains=q) | Q(requester__last_name__icontains=q)
        )
    if status:
        qs = qs.filter(status=status)
    if dept_id:
        qs = qs.filter(department_id=dept_id)
    if date_from:
        qs = qs.filter(created_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(created_at__date__lte=date_to)

    qs = qs.order_by('-created_at')
    total_count = qs.count()

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))

    return render(request, 'payments/list.html', {
        'page_obj': page_obj,
        'total_count': total_count,
        'status_choices': PaymentRequest.Status.choices,
        'departments': Department.objects.filter(is_active=True),
    })


@login_required
def my_requests_view(request):
    qs = PaymentRequest.objects.filter(requester=request.user).select_related('department', 'category').order_by('-created_at')
    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page'))
    return render(request, 'payments/list.html', {
        'page_obj': page_obj,
        'total_count': qs.count(),
        'status_choices': PaymentRequest.Status.choices,
        'departments': Department.objects.filter(is_active=True),
        'my_requests': True,
    })


@login_required
def payment_create_view(request):
    if not (request.user.is_requester or request.user.is_system_admin):
        messages.error(request, 'Only requesters can create payment requests.')
        return redirect('dashboard:home')

    if request.method == 'POST':
        form = PaymentRequestForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.requester = request.user

            # Ensure department is always set
            if not payment.department_id:
                if request.user.department:
                    payment.department = request.user.department
                else:
                    dept_id = request.POST.get('department')
                    if dept_id:
                        try:
                            payment.department = Department.objects.get(pk=dept_id)
                        except Exception:
                            pass

            if not payment.department_id:
                form.add_error('department', 'Department is required.')
                return render(request, 'payments/create.html', {'form': form})

            payment.save()

            files = request.FILES.getlist('documents')
            for f in files:
                RequestDocument.objects.create(
                    request=payment, title=f.name, file=f,
                    file_size=f.size, uploaded_by=request.user,
                )

            try:
                AuditLog.log(
                    user=request.user,
                    event_type=AuditLog.EventType.CREATE,
                    model_name='PaymentRequest',
                    object_id=payment.id,
                    object_repr=str(payment),
                    ip_address=get_client_ip(request),
                )
            except Exception:
                pass

            action = request.POST.get('action', 'draft')
            if action == 'submit':
                try:
                    ApprovalService.submit_request(payment, request.user, request=request)
                    messages.success(request, f'Request {payment.request_number} submitted for approval.')
                except Exception as e:
                    messages.warning(request, f'Request saved but could not submit: {e}')
            else:
                messages.success(request, f'Request {payment.request_number} saved as draft.')

            return redirect('payments:detail', pk=payment.pk)
    else:
        form = PaymentRequestForm(user=request.user)

    return render(request, 'payments/create.html', {'form': form})


@login_required
def payment_edit_view(request, pk):
    payment = get_object_or_404(PaymentRequest, pk=pk)

    if payment.requester != request.user and not request.user.is_system_admin:
        messages.error(request, 'You cannot edit this request.')
        return redirect('payments:detail', pk=pk)

    if payment.status not in [PaymentRequest.Status.DRAFT, PaymentRequest.Status.REJECTED_L1, PaymentRequest.Status.REJECTED_L2]:
        messages.error(request, 'This request cannot be edited in its current state.')
        return redirect('payments:detail', pk=pk)

    if request.method == 'POST':
        form = PaymentRequestForm(request.POST, instance=payment, user=request.user)
        if form.is_valid():
            form.save()
            files = request.FILES.getlist('documents')
            for f in files:
                RequestDocument.objects.create(
                    request=payment, title=f.name, file=f,
                    file_size=f.size, uploaded_by=request.user,
                )
            messages.success(request, 'Request updated successfully.')
            return redirect('payments:detail', pk=pk)
    else:
        form = PaymentRequestForm(instance=payment, user=request.user)

    return render(request, 'payments/create.html', {'form': form, 'edit': True, 'payment': payment})


@login_required
def payment_detail_view(request, pk):
    payment = get_object_or_404(
        PaymentRequest.objects.select_related(
            'requester', 'department', 'category',
            'reviewer1', 'reviewer2', 'finance_officer', 'auditor'
        ).prefetch_related('documents', 'comments__author', 'approval_actions__actor'),
        pk=pk
    )

    # Access control
    user = request.user
    can_view = (
        user.is_system_admin or user.is_auditor or
        payment.requester == user or
        (user.is_reviewer1 and payment.status == PaymentRequest.Status.SUBMITTED) or
        (user.is_reviewer2 and payment.status == PaymentRequest.Status.APPROVED_L1) or
        user.is_finance or
        payment.reviewer1 == user or payment.reviewer2 == user
    )
    if not can_view:
        messages.error(request, 'You do not have access to this request.')
        return redirect('dashboard:home')

    # Build approval chain for display
    s = payment.status
    approval_chain = [
        ('Reviewer 1 (Supervisor)', payment.reviewer1,
         s not in ['DRAFT', 'SUBMITTED'], 'check-circle-fill' if 'REJECTED_L1' not in s else 'x-circle-fill'),
        ('Reviewer 2 (Finance Controller)', payment.reviewer2,
         s in ['APPROVED_L2', 'PROCESSING', 'PAID', 'RETIREMENT_SUBMITTED', 'RETIREMENT_APPROVED', 'CLOSED'],
         'check-circle-fill'),
        ('Finance Officer', payment.finance_officer,
         s in ['PAID', 'RETIREMENT_SUBMITTED', 'RETIREMENT_APPROVED', 'CLOSED'],
         'check-circle-fill'),
        ('Auditor / Verifier', payment.auditor,
         s == 'CLOSED',
         'check-circle-fill'),
    ]

    AuditLog.log(
        user=user, event_type=AuditLog.EventType.VIEW,
        model_name='PaymentRequest', object_id=payment.id,
        object_repr=str(payment), ip_address=get_client_ip(request),
    )

    return render(request, 'payments/detail.html', {
        'payment_request': payment,
        'approval_chain': approval_chain,
    })


@login_required
@require_POST
def submit_request_view(request, pk):
    payment = get_object_or_404(PaymentRequest, pk=pk)
    try:
        remarks = request.POST.get('remarks', '')
        ApprovalService.submit_request(payment, request.user, remarks=remarks, request=request)
        messages.success(request, f'{payment.request_number} submitted for approval.')
    except Exception as e:
        messages.error(request, str(e))
    return redirect('payments:detail', pk=pk)


@login_required
@require_POST
def add_comment_view(request, pk):
    payment = get_object_or_404(PaymentRequest, pk=pk)
    content = request.POST.get('content', '').strip()
    if content:
        is_internal = bool(request.POST.get('is_internal')) and (
            request.user.is_finance or request.user.is_auditor or request.user.is_system_admin
        )
        Comment.objects.create(
            request=payment,
            author=request.user,
            content=content,
            is_internal=is_internal,
        )
        messages.success(request, 'Comment added.')
    return redirect('payments:detail', pk=pk)


@login_required
@require_POST
def submit_retirement_view(request, pk):
    payment = get_object_or_404(PaymentRequest, pk=pk)
    try:
        amount_actual = request.POST.get('amount_actual')
        remarks = request.POST.get('remarks', '')
        if not amount_actual:
            raise ValueError('Actual amount is required.')

        ApprovalService.submit_retirement(
            payment, request.user,
            amount_actual=float(amount_actual),
            remarks=remarks, request=request
        )

        # Handle receipt uploads
        receipts = request.FILES.getlist('receipts')
        for f in receipts:
            RequestDocument.objects.create(
                request=payment, title=f.name, file=f,
                file_size=f.size, uploaded_by=request.user,
                doc_type=RequestDocument.DocType.RECEIPT,
                is_retirement_doc=True,
            )

        messages.success(request, 'Retirement submitted successfully.')
    except Exception as e:
        messages.error(request, str(e))
    return redirect('payments:detail', pk=pk)


@login_required
@require_POST
def upload_doc_view(request, pk):
    payment = get_object_or_404(PaymentRequest, pk=pk)
    files = request.FILES.getlist('documents')
    if files:
        for f in files:
            RequestDocument.objects.create(
                request=payment, title=f.name, file=f,
                file_size=f.size, uploaded_by=request.user,
            )
        messages.success(request, f'{len(files)} document(s) uploaded.')
    return redirect('payments:detail', pk=pk)


# ─── EXPORT ───────────────────────────────────────────────────
@login_required
def export_excel_view(request):
    if not (request.user.is_system_admin or request.user.is_finance or request.user.is_auditor):
        messages.error(request, 'Export access denied.')
        return redirect('payments:list')

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        qs = PaymentRequest.objects.select_related('requester', 'department').order_by('-created_at')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Payment Requests'

        # Header styling
        header_fill = PatternFill(start_color='0A2463', end_color='0A2463', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        headers = [
            'Reference', 'Title', 'Department', 'Requester',
            'Amount Requested', 'Amount Approved', 'Amount Actual',
            'Status', 'Priority', 'Submitted Date', 'Paid Date', 'Closed Date'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[1].height = 22

        for row, req in enumerate(qs, 2):
            ws.cell(row=row, column=1, value=req.request_number)
            ws.cell(row=row, column=2, value=req.title)
            ws.cell(row=row, column=3, value=str(req.department))
            ws.cell(row=row, column=4, value=req.requester.full_name)
            ws.cell(row=row, column=5, value=float(req.amount_requested))
            ws.cell(row=row, column=6, value=float(req.amount_approved) if req.amount_approved else '')
            ws.cell(row=row, column=7, value=float(req.amount_actual) if req.amount_actual else '')
            ws.cell(row=row, column=8, value=req.get_status_display())
            ws.cell(row=row, column=9, value=req.get_priority_display())
            ws.cell(row=row, column=10, value=req.submitted_at.strftime('%Y-%m-%d %H:%M') if req.submitted_at else '')
            ws.cell(row=row, column=11, value=req.paid_at.strftime('%Y-%m-%d %H:%M') if req.paid_at else '')
            ws.cell(row=row, column=12, value=req.closed_at.strftime('%Y-%m-%d %H:%M') if req.closed_at else '')

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"wolecen_payment_requests_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        AuditLog.log(
            user=request.user, event_type=AuditLog.EventType.EXPORT,
            ip_address=get_client_ip(request),
        )

        return response
    except ImportError:
        messages.error(request, 'Excel export requires openpyxl. Run: pip install openpyxl')
        return redirect('payments:list')


@login_required
def export_pdf_view(request, pk):
    payment = get_object_or_404(PaymentRequest, pk=pk)

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.units import cm

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm, leftMargin=2*cm, rightMargin=2*cm)

        styles = getSampleStyleSheet()
        navy = colors.HexColor('#0A2463')
        accent = colors.HexColor('#E63946')
        light_gray = colors.HexColor('#F8FAFC')
        border_color = colors.HexColor('#E2E8F0')

        title_style = ParagraphStyle('Title', parent=styles['Normal'], fontSize=18, fontName='Helvetica-Bold', textColor=navy, spaceAfter=4)
        heading_style = ParagraphStyle('Heading', parent=styles['Normal'], fontSize=11, fontName='Helvetica-Bold', textColor=navy, spaceBefore=12, spaceAfter=4)
        body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, spaceAfter=4, leading=14)
        label_style = ParagraphStyle('Label', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#718096'), fontName='Helvetica-Bold', spaceAfter=2)

        story = []

        # Header
        story.append(Paragraph('WOLECEN ENGINEERING GROUP LIMITED', ParagraphStyle('Co', parent=styles['Normal'], fontSize=14, fontName='Helvetica-Bold', textColor=navy)))
        story.append(Paragraph('Payment Request & Retirement System', ParagraphStyle('Sub', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#718096'), spaceAfter=4)))
        story.append(HRFlowable(width='100%', thickness=2, color=navy, spaceAfter=12))

        # Request reference
        story.append(Paragraph(f'PAYMENT REQUEST — {payment.request_number}', title_style))
        story.append(Paragraph(payment.title, ParagraphStyle('T2', parent=styles['Normal'], fontSize=14, textColor=colors.HexColor('#2D3748'), spaceAfter=8)))

        # Details table
        detail_data = [
            ['Status', payment.get_status_display(), 'Priority', payment.get_priority_display()],
            ['Department', str(payment.department), 'Category', str(payment.category or '—')],
            ['Requester', payment.requester.full_name, 'Date Submitted', payment.submitted_at.strftime('%d %B %Y') if payment.submitted_at else '—'],
            ['Amount Requested', f"TZS {float(payment.amount_requested):,.0f}", 'Amount Approved', f"TZS {float(payment.amount_approved):,.0f}" if payment.amount_approved else '—'],
        ]
        if payment.amount_actual:
            detail_data.append(['Amount Actual', f"TZS {float(payment.amount_actual):,.0f}", 'Variance', f"TZS {float(payment.amount_variance):,.0f}" if payment.amount_variance else '—'])

        tbl = Table(detail_data, colWidths=[3.5*cm, 6*cm, 3.5*cm, 4*cm])
        tbl.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), light_gray),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#EDF2F7')),
            ('BACKGROUND', (2, 0), (2, -1), colors.HexColor('#EDF2F7')),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('PADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        story.append(tbl)
        story.append(Spacer(1, 0.4*cm))

        # Purpose
        story.append(Paragraph('Purpose & Justification', heading_style))
        story.append(Paragraph(payment.purpose, body_style))

        # Beneficiary
        if payment.beneficiary_name:
            story.append(Paragraph('Beneficiary Details', heading_style))
            bene_data = [
                ['Name', payment.beneficiary_name],
                ['Bank', payment.beneficiary_bank or '—'],
                ['Account', payment.beneficiary_account or '—'],
            ]
            bt = Table(bene_data, colWidths=[4*cm, 13*cm])
            bt.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, border_color),
                ('PADDING', (0, 0), (-1, -1), 6),
                ('BACKGROUND', (0, 0), (0, -1), light_gray),
            ]))
            story.append(bt)

        # Approval trail
        story.append(Paragraph('Approval Trail', heading_style))
        trail_data = [['Action', 'By', 'Date', 'Remarks']]
        for action in payment.approval_actions.all():
            trail_data.append([
                action.get_action_display(),
                action.actor.full_name,
                action.timestamp.strftime('%d/%m/%Y %H:%M'),
                (action.remarks or '—')[:60],
            ])
        at = Table(trail_data, colWidths=[4.5*cm, 4.5*cm, 3.5*cm, 5*cm])
        at.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), navy),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, border_color),
            ('PADDING', (0, 0), (-1, -1), 5),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light_gray]),
        ]))
        story.append(at)

        # Signature block
        story.append(Spacer(1, 1*cm))
        story.append(HRFlowable(width='100%', thickness=0.5, color=border_color))
        story.append(Spacer(1, 0.3*cm))
        sig_data = [['Requester Signature', 'Reviewer 1 Signature', 'Finance Controller', 'Finance Officer']]
        sig_data.append(['', '', '', ''])
        sig_data.append(['Date: ___________', 'Date: ___________', 'Date: ___________', 'Date: ___________'])
        st = Table(sig_data, colWidths=[4.25*cm, 4.25*cm, 4.25*cm, 4.25*cm])
        st.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 1), (-1, 1), 24),
        ]))
        story.append(st)

        # Footer
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(
            f'Generated by Wolecen EGL Payment System — {timezone.now().strftime("%d %B %Y %H:%M")} — Confidential',
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=7, textColor=colors.HexColor('#A0AEC0'), alignment=1)
        ))

        doc.build(story)
        buffer.seek(0)

        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{payment.request_number}.pdf"'
        return response

    except ImportError:
        messages.error(request, 'PDF export requires reportlab. Run: pip install reportlab')
        return redirect('payments:detail', pk=pk)


@login_required
def print_view(request, pk):
    payment = get_object_or_404(
        PaymentRequest.objects.select_related('requester', 'department', 'reviewer1', 'reviewer2', 'finance_officer', 'auditor').prefetch_related('documents', 'approval_actions__actor'),
        pk=pk
    )
    return render(request, 'payments/print.html', {'payment_request': payment})


# ─── CONTEXT PROCESSOR ────────────────────────────────────────
def notification_count(request):
    if not request.user.is_authenticated:
        return {}

    counts = {}
    user = request.user

    if user.can_approve_level1:
        counts['pending_l1_count'] = PaymentRequest.objects.filter(status=PaymentRequest.Status.SUBMITTED).count()
    if user.can_approve_level2:
        counts['pending_l2_count'] = PaymentRequest.objects.filter(status=PaymentRequest.Status.APPROVED_L1).count()
    if user.can_process_payment:
        counts['finance_count'] = PaymentRequest.objects.filter(
            status__in=[PaymentRequest.Status.APPROVED_L2, PaymentRequest.Status.PROCESSING]
        ).count()
    if user.can_verify_retirement:
        counts['retirement_count'] = PaymentRequest.objects.filter(status=PaymentRequest.Status.RETIREMENT_SUBMITTED).count()

    total = sum(counts.values())
    counts['notification_count'] = total

    return counts
